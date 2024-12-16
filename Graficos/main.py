import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
import argparse
import os
import sys

def extrair_tempo_execucao(linha):
    try:
        return float(linha.strip()) 
    except ValueError:
        return None

def extrair_memoria(linha):
    try:
        return int(linha.strip())
    except ValueError:
        return None

parser = argparse.ArgumentParser()
parser.add_argument("caminho")

args = parser.parse_args()

if not os.path.exists(args.caminho):
    print('Erro no caminho')
    sys.exit()


with open(args.caminho, 'r') as f:
    linhas = f.readlines()

tempos_execucao = []
usos_memoria = []

for i in range(0, len(linhas), 2):
    tempo_execucao = extrair_tempo_execucao(linhas[i])
    uso_memoria = extrair_memoria(linhas[i+1]) if i+1 < len(linhas) else None

    if tempo_execucao is not None and uso_memoria is not None:
        tempos_execucao.append(tempo_execucao)
        usos_memoria.append(uso_memoria)

# Criar DataFrame
df = pd.DataFrame({
    "Tempo de Execução (s)": tempos_execucao,
    "Uso de Memória (KB)": usos_memoria
})

# Salvar DataFrame
excel_path = f'Analytics/{args.caminho}.xlsx'
df.to_excel(excel_path, index=False)

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

font_bold = Font(bold=True, name='Cambria', size=11)

media_tempo = df["Tempo de Execução (s)"].mean()
mediana_tempo = df["Tempo de Execução (s)"].median()
media_memoria = df["Uso de Memória (KB)"].mean()
mediana_memoria = df["Uso de Memória (KB)"].median()
ws.append(["Média Tempo de Execução (s)", media_tempo])
ws.append(["Mediana Tempo de Execução (s)", mediana_tempo])
ws.append(["Média Uso de Memória (KB)", media_memoria])
ws.append(["Mediana Uso de Memória (KB)", mediana_memoria])
ws["A" + str(len(tempos_execucao) + 2)].font = font_bold
ws["A" + str(len(tempos_execucao) + 3)].font = font_bold
ws["A" + str(len(tempos_execucao) + 4)].font = font_bold
ws["A" + str(len(tempos_execucao) + 5)].font = font_bold

# Ajusta largura das colunas pra evitar sobreposição
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # letra da coluna
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)  # Ajusta largura da coluna
    ws.column_dimensions[column].width = adjusted_width

# Adicionar um gráfico de barras
chart = BarChart()
chart.title = "Tempo de Execução e Uso de Memória"
chart.style = 10
chart.x_axis.title = 'Tempo'
chart.y_axis.title = 'Memória'

# Definir os dados do gráfico
data = Reference(ws, min_col=1, min_row=2, max_col=2, max_row=len(tempos_execucao)+1)
chart.add_data(data, titles_from_data=True)
categories = Reference(ws, min_col=1, min_row=2, max_row=len(tempos_execucao)+1)
chart.set_categories(categories)

# Posiciona o gráfico na planilha
ws.add_chart(chart, "E1")

wb.save(excel_path)
